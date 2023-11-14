"""
    本段代码实现：
        1.向Vrprofile2网站批量提交fna数据，获取用于回溯的jobID（网站运行需要一定时间）
        2.基于jobID，将网站运行结果下载到本地
"""
import os
import re
import time
from openpyxl import Workbook
import requests
from concurrent.futures import ThreadPoolExecutor


def put_file(x, filename, id):
    print(x, filename, id)
    headers = {
        'Host': 'tool2-mml.sjtu.edu.cn',
        'Cache-Control': 'max-age=0',
        'sec-ch-ua': '"Chromium";v="104", " Not A;Brand";v="99", "Microsoft Edge";v="104"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
        'Upgrade-Insecure-Requests': '1',
        'Origin': 'https://tool2-mml.sjtu.edu.cn',
        'Content-Type': 'multipart/form-data; boundary=----WebKitFormBoundaryPnrYhj4aC8y9CD0E',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.5112.102 Safari/537.36 Edg/104.0.1293.70',
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'Sec-Fetch-Site': 'same-origin',
        'Sec-Fetch-Mode': 'navigate',
        'Sec-Fetch-User': '?1',
        'Sec-Fetch-Dest': 'document',
        'Referer': 'https://tool2-mml.sjtu.edu.cn/VRprofile/VRprofile.php',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
    }
    f = open(fr'Bacillus thuringiensis/{filename}', 'rb').read().decode('utf-8')    # 读取文件，根据文件内容判断是质粒还是染色体
    try:
        if f.index('plasmid'):
            dna_type = 'Plasmid'
    except Exception as e:
        dna_type = 'Chromosome'
    data = '------WebKitFormBoundaryPnrYhj4aC8y9CD0E\nContent-Disposition: form-data; name="inputFile"; filename="{}"\nContent-Type: application/octet-stream\n\n{}------WebKitFormBoundaryPnrYhj4aC8y9CD0E\nContent-Disposition: form-data; name="optionsRadios"\n\n{}\n------WebKitFormBoundaryPnrYhj4aC8y9CD0E\nContent-Disposition: form-data; name="entry"\n\n{}\n------WebKitFormBoundaryPnrYhj4aC8y9CD0E--\n'.format(
        filename, f, dna_type, id)
    response = requests.post('https://tool2-mml.sjtu.edu.cn/cgi-bin/VRprofile/VRprofile1.cgi', headers=headers,
                             data=data)     # 向网站上传文件
    # 将文件名和对应的网址写到excel中
    ws[f'A{x}'].value = filename    # excel的A列写filename
    ws[f'B{x}'].value = f'https://tool2-mml.sjtu.edu.cn/VRprofile/angular1.php?ty=c&job={id}'   # excel的B列写网址
    print(response.status_code, f'上传成功 {dna_type} 网址 https://tool2-mml.sjtu.edu.cn/VRprofile/angular1.php?ty=c&job={id}')


def dispose_page(x, y):
    headers = {
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,image/apng,*/*;q=0.8,application/signed-exchange;v=b3;q=0.9',
        'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6',
        'Cache-Control': 'no-cache',
        'Connection': 'keep-alive',
        'Pragma': 'no-cache',
        'Referer': 'https://tool2-mml.sjtu.edu.cn/VRprofile/VRprofile.php',
        'Sec-Fetch-Dest': 'document',
        'Sec-Fetch-Mode': 'navigate',
        'Sec-Fetch-Site': 'same-origin',
        'Sec-Fetch-User': '?1',
        'Upgrade-Insecure-Requests': '1',
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/104.0.5112.102 Safari/537.36 Edg/104.0.1293.70',
        'sec-ch-ua': '"Chromium";v="104", " Not A;Brand";v="99", "Microsoft Edge";v="104"',
        'sec-ch-ua-mobile': '?0',
        'sec-ch-ua-platform': '"Windows"',
    }
    response = requests.get('https://tool2-mml.sjtu.edu.cn/VRprofile/VRprofile.php', headers=headers)
    if response.status_code == 200:  # 网页解析成功，进行下一步
        id = re.findall('randomString = "(.+)";', response.text)[0]
        put_file(x, y, id)    # 提交文件
    else:
        print("失败---------------------------------------")


def file_submit():
    wb = Workbook()
    ws = wb.active
    with ThreadPoolExecutor(8) as tb:
        for x, y in enumerate(os.listdir('Bacillus thuringiensis')):
            x = x + 1
            tb.submit(dispose_page, x, y)
    wb.save('完成.xlsx')



def fileDown(jobID, file_type):
    url = 'https://tool2-mml.sjtu.edu.cn/VRprofile/down.php?ac='+str(file_type)+'&job='+jobID
    try:
        r = requests.get(url)
        r.raise_for_status()
        r.encoding = r.apparent_encoding
        filename = str(file_type)+'_'+str(jobID)
        with open('VRprofile_Download/'+filename+'.txt','w')as file:
            file.write(r.text)
        file.close()
    except:
        print(str(jobID)+'的'+str(file_type)+'网页爬取失败')
        print(str(url))


def file_down(input_file):
    f = open(input_file)
    lines = f.readlines()
    m = 0
    for job_info in lines:
        m = m + 1
        jobID, job_type = [x for x in job_info.split('\t')]
        fileDown(jobID, job_type)